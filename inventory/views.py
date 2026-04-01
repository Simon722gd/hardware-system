import base64
import json
import requests
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from .models import Category, Product, Sale, Supplier
from .forms import CategoryForm, ProductForm, SupplierForm, SaleForm, UserForm

# Import for PDF export (optional)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# Import for Excel export (optional)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from datetime import datetime

def normalize_mpesa_phone(phone):
    if not phone:
        return None
    normalized = phone.strip().replace(' ', '').replace('+', '')
    if normalized.startswith('07') and len(normalized) == 10 and normalized.isdigit():
        return '254' + normalized[1:]
    if normalized.startswith('2547') and len(normalized) == 12 and normalized.isdigit():
        return normalized
    return None

def get_mpesa_access_token():
    consumer_key = getattr(settings, 'MPESA_CONSUMER_KEY', '').strip()
    consumer_secret = getattr(settings, 'MPESA_CONSUMER_SECRET', '').strip()
    if not consumer_key or not consumer_secret:
        raise ValueError('MPesa consumer credentials are not configured.')

    token_url = 'https://sandbox.safaricom.co.ke/oauth/v1/generate?grant_type=client_credentials'
    response = requests.get(token_url, auth=(consumer_key, consumer_secret), timeout=30)
    response.raise_for_status()
    token_data = response.json()
    access_token = token_data.get('access_token')
    print('MPesa access token response:', token_data)
    if not access_token:
        raise ValueError('Unable to obtain MPesa access token.')
    print('MPesa access token:', access_token)
    return access_token


@login_required(login_url='accounts:login')
def mpesa_token_test(request):
    try:
        token = get_mpesa_access_token()
    except ValueError as exc:
        return HttpResponse(f'M-Pesa credentials not set: {exc}', status=500)
    except requests.RequestException as exc:
        return HttpResponse(f'Error fetching M-Pesa token: {exc}', status=500)

    print('M-Pesa access token:', token)
    return HttpResponse('Access token generated')


def build_mpesa_password(timestamp):
    shortcode = getattr(settings, 'MPESA_SHORTCODE', '')
    passkey = getattr(settings, 'MPESA_PASSKEY', '')
    if not shortcode or not passkey:
        raise ValueError('MPesa shortcode or passkey is not configured.')
    raw_password = f"{shortcode}{passkey}{timestamp}"
    return base64.b64encode(raw_password.encode()).decode()


def initiate_mpesa_stk_push(phone_number, amount, account_reference, transaction_desc, callback_url):
    access_token = get_mpesa_access_token()
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    password = build_mpesa_password(timestamp)
    shortcode = getattr(settings, 'MPESA_SHORTCODE', '')
    endpoint = 'https://sandbox.safaricom.co.ke/mpesa/stkpush/v1/processrequest'
    payload = {
        'BusinessShortCode': shortcode,
        'Password': password,
        'Timestamp': timestamp,
        'TransactionType': 'CustomerPayBillOnline',
        'Amount': int(round(float(amount))),
        'PartyA': phone_number,
        'PartyB': shortcode,
        'PhoneNumber': phone_number,
        'CallBackURL': callback_url,
        'AccountReference': account_reference,
        'TransactionDesc': transaction_desc,
    }
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json',
    }
    print('MPesa STK push request payload:', payload)
    response = requests.post(endpoint, json=payload, headers=headers, timeout=30)
    print('MPesa STK push HTTP status:', response.status_code)
    response.raise_for_status()
    response_data = response.json()
    print('MPesa STK push response:', response_data)
    return response_data

@login_required(login_url='accounts:login')
def dashboard(request):
    total_categories = Category.objects.count()
    total_products = Product.objects.count()
    total_sales = Sale.objects.count()
    low_stock_items = Product.objects.filter(quantity__lt=10).count()
    recent_activity = Sale.objects.order_by('-date')[:5]

    # Chart data
    products_per_category = Category.objects.annotate(product_count=Count('products')).values('name', 'product_count')
    sales_over_time = Sale.objects.annotate(month=TruncMonth('date')).values('month').annotate(sale_count=Count('id')).order_by('month')

    cat_labels = [item['name'] for item in products_per_category]
    cat_data = [item['product_count'] for item in products_per_category]
    
    # Sales labels (e.g., month names)
    sales_labels = [item['month'].strftime('%b %Y') for item in sales_over_time]
    sales_data = [item['sale_count'] for item in sales_over_time]

    context = {
        'total_categories': total_categories,
        'total_products': total_products,
        'total_sales': total_sales,
        'low_stock_items': low_stock_items,
        'recent_activity': recent_activity,
        'cat_labels': json.dumps(cat_labels),
        'cat_data': json.dumps(cat_data),
        'sales_labels': json.dumps(sales_labels),
        'sales_data': json.dumps(sales_data),
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required(login_url='accounts:login')
def products_view(request):
    products = Product.objects.all()
    return render(request, 'inventory/products.html', {'products': products})

# Low Stock Products
@login_required(login_url='accounts:login')
def low_stock_products(request):
    low_stock_items = Product.objects.filter(quantity__lt=10)
    context = {
        'products': low_stock_items,
        'total_low_stock': low_stock_items.count(),
    }
    return render(request, 'inventory/low_stock_products.html', context)

# Search Functionality
@login_required(login_url='accounts:login')
def search(request):
    query = request.GET.get('q', '').strip()
    search_type = request.GET.get('type', 'all').strip()
    
    results = {
        'products': [],
        'categories': [],
        'suppliers': [],
        'query': query,
        'search_type': search_type,
    }
    
    if query:
        if search_type in ['all', 'products']:
            results['products'] = Product.objects.filter(
                Q(name__icontains=query) | Q(sku__icontains=query)
            )
        if search_type in ['all', 'categories']:
            results['categories'] = Category.objects.filter(name__icontains=query)
        if search_type in ['all', 'suppliers']:
            results['suppliers'] = Supplier.objects.filter(name__icontains=query)
    
    return render(request, 'inventory/search_results.html', results)


# Category CRUD
@login_required(login_url='accounts:login')
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'inventory/category_list.html', {'categories': categories})

@login_required(login_url='accounts:login')
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category created successfully.')
            return redirect('inventory:category_list')
    else:
        form = CategoryForm()
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Create Category'})

@login_required(login_url='accounts:login')
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully.')
            return redirect('inventory:category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Update Category'})

@login_required(login_url='accounts:login')
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted successfully.')
        return redirect('inventory:category_list')
    return render(request, 'inventory/delete_confirm.html', {'object': category})

# Product CRUD
@login_required(login_url='accounts:login')
def product_list(request):
    products = Product.objects.all()
    return render(request, 'inventory/product_list.html', {'products': products})

@login_required(login_url='accounts:login')
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product created successfully.')
            return redirect('inventory:product_list')
    else:
        form = ProductForm()
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Create Product'})

@login_required(login_url='accounts:login')
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully.')
            return redirect('inventory:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Update Product'})

@login_required(login_url='accounts:login')
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully.')
        return redirect('inventory:product_list')
    return render(request, 'inventory/delete_confirm.html', {'object': product})

# Supplier CRUD
@login_required(login_url='accounts:login')
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers})

@login_required(login_url='accounts:login')
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier created successfully.')
            return redirect('inventory:supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Create Supplier'})

@login_required(login_url='accounts:login')
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier updated successfully.')
            return redirect('inventory:supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Update Supplier'})

@login_required(login_url='accounts:login')
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, 'Supplier deleted successfully.')
        return redirect('inventory:supplier_list')
    return render(request, 'inventory/delete_confirm.html', {'object': supplier})

# Sales CRUD
@login_required(login_url='accounts:login')
def sale_list(request):
    sales = Sale.objects.select_related('product').all().order_by('-date')
    return render(request, 'inventory/sale_list.html', {'sales': sales})

@login_required(login_url='accounts:login')
def sale_create(request):
    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            sale.amount_paid = sale.amount_paid or 0
            product = sale.product
            if sale.quantity > product.quantity:
                messages.error(request, 'Insufficient stock for this product.')
                return redirect('inventory:sale_create')

            sale.total_price = product.price * sale.quantity
            sale.balance = max(0, sale.total_price - sale.amount_paid)
            sale.status = 'Paid' if sale.amount_paid >= sale.total_price else 'Pending'

            if sale.payment_method == 'Cash' and not sale.amount_paid:
                sale.amount_paid = sale.total_price
                sale.balance = 0
                sale.status = 'Paid'

            payment_action = request.POST.get('payment_action', '')
            if payment_action == 'stk_push' and sale.payment_method == 'M-Pesa':
                normalized_phone = normalize_mpesa_phone(sale.phone_number)
                if not normalized_phone:
                    form.add_error('phone_number', 'Phone number must be in 2547XXXXXXXX format or start with 07.')
                else:
                    sale.phone_number = normalized_phone
                    callback_url = getattr(settings, 'MPESA_CALLBACK_URL', '') or request.build_absolute_uri(reverse('mpesa_callback'))
                    try:
                        mpesa_response = initiate_mpesa_stk_push(
                            phone_number=normalized_phone,
                            amount=sale.total_price,
                            account_reference='HardwareSystem',
                            transaction_desc='Payment',
                            callback_url=callback_url
                        )
                    except requests.RequestException as exc:
                        messages.error(request, f'M-Pesa request failed: {exc}')
                    except ValueError as exc:
                        messages.error(request, f'M-Pesa configuration issue: {exc}')
                    else:
                        response_code = str(mpesa_response.get('ResponseCode') or mpesa_response.get('responseCode', ''))
                        response_description = mpesa_response.get('ResponseDescription') or mpesa_response.get('errorMessage') or 'No response description.'
                        if response_code == '0':
                            print('STK Push sent successfully')
                            sale.notes = json.dumps({
                                'mpesa_request': {
                                    'MerchantRequestID': mpesa_response.get('MerchantRequestID'),
                                    'CheckoutRequestID': mpesa_response.get('CheckoutRequestID'),
                                    'ResponseCode': response_code,
                                    'ResponseDescription': response_description,
                                    'phone': normalized_phone,
                                }
                            })
                            sale.save()
                            product.quantity -= sale.quantity
                            product.save()
                            messages.success(request, 'M-Pesa payment request sent successfully. Complete payment on customer phone.')
                            return redirect('inventory:sale_receipt', pk=sale.pk)
                        print('STK Push error:', response_description)
                        messages.error(request, f'M-Pesa request failed: {response_description}')

            if payment_action != 'stk_push':
                sale.save()
                product.quantity -= sale.quantity
                product.save()
                messages.success(request, 'Sale created successfully.')
                return redirect('inventory:sale_receipt', pk=sale.pk)
    else:
        form = SaleForm()
        
    # Get product prices for dynamic frontend calculation
    products = Product.objects.all()
    product_prices = {product.id: float(product.price) for product in products}
    
    return render(request, 'inventory/sale_form.html', {
        'form': form, 
        'title': 'Create Sale',
        'product_prices_json': json.dumps(product_prices)
    })

@login_required(login_url='accounts:login')
def sale_receipt(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'inventory/sale_receipt.html', {'sale': sale})

@login_required(login_url='accounts:login')
def mpesa_payment(request):
    error = None
    phone = ''

    if request.method == 'POST':
        phone = (request.POST.get('phone') or '').strip()
        if not phone:
            error = 'Phone number is required.'
        elif not phone.isdigit() or not phone.startswith('254') or len(phone) != 12:
            error = 'Phone number must start with 254 and contain 12 digits.'
        else:
            print('Phone Number:', phone)
            return HttpResponse(f'Payment request sent to {phone}')

    return render(request, 'inventory/mpesa_payment.html', {
        'error': error,
        'phone': phone,
    })

@csrf_exempt
def mpesa_callback(request):
    if request.method != 'POST':
        return HttpResponseBadRequest('Invalid callback method.')

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except ValueError:
        return HttpResponseBadRequest('Invalid JSON payload.')

    print('MPesa callback received:', json.dumps(payload, indent=2))
    callback_data = payload.get('Body', {}).get('stkCallback', {})
    checkout_request_id = callback_data.get('CheckoutRequestID')
    merchant_request_id = callback_data.get('MerchantRequestID')
    result_code = callback_data.get('ResultCode')
    result_desc = callback_data.get('ResultDesc')
    print('MPesa callback result:', result_code, result_desc)

    sale = None
    if checkout_request_id:
        sale = Sale.objects.filter(notes__contains=checkout_request_id).first()
    if not sale and merchant_request_id:
        sale = Sale.objects.filter(notes__contains=merchant_request_id).first()

    if sale:
        sale.notes = (sale.notes or '') + '\n' + json.dumps({'mpesa_callback': callback_data})
        if result_code == 0:
            sale.status = 'Paid'
            sale.amount_paid = sale.total_price
            sale.balance = 0
        sale.save()

    return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})

@login_required(login_url='accounts:login')
def payments(request):
    payments = Sale.objects.select_related('product').order_by('-date')
    payment_method = request.GET.get('payment_method')
    status = request.GET.get('status')

    if payment_method in ['Cash', 'M-Pesa']:
        payments = payments.filter(payment_method=payment_method)
    if status in ['Paid', 'Pending']:
        payments = payments.filter(status=status)

    return render(request, 'inventory/payments.html', {
        'title': 'Payments',
        'payments': payments,
    })

@login_required(login_url='accounts:login')
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    sales = product.sales.order_by('-date')
    total_sold = sales.aggregate(total=Sum('quantity'))['total'] or 0
    total_revenue = sales.aggregate(total=Sum('total_price'))['total'] or 0
    return render(request, 'inventory/product_detail.html', {
        'product': product,
        'sales': sales,
        'total_sold': total_sold,
        'total_revenue': total_revenue,
    })

# Reports
@login_required(login_url='accounts:login')
def reports(request):
    selected_product_id = request.GET.get('product')
    products = Product.objects.all()
    sales = Sale.objects.select_related('product').all().order_by('-date')

    if selected_product_id:
        sales = sales.filter(product_id=selected_product_id)

    total_sales_count = sales.count()
    total_revenue = sales.aggregate(total=Sum('total_price'))['total'] or 0
    monthly_sales = sales.annotate(month=TruncMonth('date')).values('month').annotate(count=Count('id'), revenue=Sum('total_price')).order_by('month')

    months = [item['month'].strftime('%b %Y') for item in monthly_sales]
    revenues = [float(item['revenue'] or 0) for item in monthly_sales]

    return render(request, 'inventory/reports.html', {
        'total_sales_count': total_sales_count,
        'total_revenue': total_revenue,
        'months_json': json.dumps(months),
        'revenues_json': json.dumps(revenues),
        'products': products,
        'selected_product_id': int(selected_product_id) if selected_product_id else None,
        'sales_history': sales,
    })

# Users
@login_required(login_url='accounts:login')
def user_list(request):
    users = User.objects.all()
    return render(request, 'inventory/user_list.html', {'users': users})

@login_required(login_url='accounts:login')
def user_add(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            # Check if user already exists
            username = form.cleaned_data.get('username')
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
            else:
                User.objects.create_user(
                    username=username,
                    email=form.cleaned_data.get('email'),
                    password=form.cleaned_data.get('password')
                )
                messages.success(request, 'User created successfully.')
                return redirect('inventory:user_list')
    else:
        form = UserForm()
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Add System User'})

# Export Functionality
@login_required(login_url='accounts:login')
def export_sales_pdf(request):
    """Export sales data to PDF"""
    if not HAS_REPORTLAB:
        messages.error(request, 'PDF export is not available. Please install reportlab: pip install reportlab')
        return redirect('inventory:reports')
    
    sales = Sale.objects.all().order_by('-date')
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2563EB'),
        spaceAfter=30,
        alignment=1,  # Center
    )
    
    title = Paragraph(f"Sales Report - {datetime.now().strftime('%Y-%m-%d')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Create table data
    table_data = [['Product', 'Quantity', 'Total Price (KES)', 'Date', 'Notes']]
    for sale in sales:
        table_data.append([
            sale.product.name,
            str(sale.quantity),
            f"{sale.total_price:,.2f}",
            sale.date.strftime('%Y-%m-%d %H:%M'),
            sale.notes or '-'
        ])
    
    # Create table
    table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1.8*inch, 1.2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.5 * inch))
    
    # Add summary
    total_revenue = Sale.objects.aggregate(total=Sum('total_price'))['total'] or 0
    summary_text = f"<b>Total Sales:</b> {sales.count()} | <b>Total Revenue:</b> KES {total_revenue:,.2f}"
    elements.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    return response

@login_required(login_url='accounts:login')
def export_products_excel(request):
    """Export products list to Excel"""
    if not HAS_OPENPYXL:
        messages.error(request, 'openpyxl library is not installed. Please install it to use Excel export.')
        return redirect('inventory:products')
    
    products = Product.objects.all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Add headers
    headers = ['Product Name', 'Category', 'Supplier', 'Price (KES)', 'Quantity', 'SKU', 'Created Date']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for product in products:
        ws.append([
            product.name,
            product.category.name,
            product.supplier.name if product.supplier else '-',
            float(product.price),
            product.quantity,
            product.sku or '-',
            product.created_at.strftime('%Y-%m-%d'),
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

